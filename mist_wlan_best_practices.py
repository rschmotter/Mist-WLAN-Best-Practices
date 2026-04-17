#!/usr/bin/env python3
"""
Juniper Mist WLAN Best Practices Automation Script
====================================================
Checks and optionally remediates WLAN SSID configuration against
Juniper Mist best practices for a given Mist Org.

Best Practices Checked:
  1. ARP Filtering
  2. Multicast/Broadcast Filtering (limit_bcast)
  3. Allow IPv6 Clients (allow_ipv6_ndp)
  4. Data Rates (no legacy 802.11b rates)
  5. 802.11r Fast Transition (for WPA2/WPA3 Enterprise)
  6. Duplicate WLAN Names per Site

Author : Generated for Mist WLAN Best Practices Project
Version: 1.0.0
"""

import sys
import time
import json
import getpass
import datetime
import argparse
import os
import re
from collections import defaultdict

# ---------------------------------------------------------------------------
# Attempt to import optional dependencies; guide user if missing
# ---------------------------------------------------------------------------
try:
    import requests
except ImportError:
    sys.exit("ERROR: 'requests' library not found.  Run: pip install requests")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("WARNING: 'openpyxl' not found – Excel export disabled.  Run: pip install openpyxl")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SCRIPT_VERSION = "1.0.0"
API_RATE_LIMIT  = 2000        # max API calls per hour
API_CALL_WINDOW = 3600        # seconds in an hour
RATE_LIMIT_BUFFER = 50        # reserve calls for safety

MIST_CLOUDS = {
    "1": {"name": "Global 01 (US)",   "base": "https://api.mist.com"},
    "2": {"name": "Global 02",        "base": "https://api.gc1.mist.com"},
    "3": {"name": "Europe (EU)",      "base": "https://api.eu.mist.com"},
    "4": {"name": "APAC (AC2)",       "base": "https://api.ac2.mist.com"},
    "5": {"name": "APAC (AC5)",       "base": "https://api.ac5.mist.com"},
    "6": {"name": "Canada (CA)",      "base": "https://api.ca.mist.com"},
}

# Enterprise auth types that support 802.11r
ENTERPRISE_AUTH_TYPES = {"eap", "eap-reauth", "dot1x_eap", "dot1x_cert", "dot1x"}

# Best-practice field settings (applied when remediating)
BP_REMEDIATION = {
    "arp_filter":      True,
    "limit_bcast":     True,
    "allow_ipv6_ndp":  True,
    # data_rates: handled separately
    # 802.11r:          handled separately (auth.disable_ft = False)
}

# ---------------------------------------------------------------------------
# Logging helpers
# ---------------------------------------------------------------------------
LOG_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(LOG_DIR, exist_ok=True)

_run_ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE  = os.path.join(LOG_DIR, f"mist_bp_{_run_ts}.log")
DEBUG_LOG = os.path.join(LOG_DIR, f"mist_bp_debug_{_run_ts}.log")

_t0 = time.time()

def _ts():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _elapsed():
    s = int(time.time() - _t0)
    return f"{s//60:02d}m{s%60:02d}s"

def log(msg, level="INFO", console=True):
    line = f"[{_ts()}] [{level}] {msg}"
    with open(LOG_FILE, "a") as f:
        f.write(line + "\n")
    if console:
        color = {"INFO": "\033[0m", "OK": "\033[92m", "WARN": "\033[93m",
                 "ERROR": "\033[91m", "DEBUG": "\033[94m", "HEAD": "\033[1m"}.get(level, "\033[0m")
        print(f"{color}{line}\033[0m")

def debug(msg):
    line = f"[{_ts()}] [DEBUG] {msg}"
    with open(DEBUG_LOG, "a") as f:
        f.write(line + "\n")

def section(title):
    bar = "=" * 70
    log(f"\n{bar}\n  {title}\n{bar}", level="HEAD")

def progress(current, total, label=""):
    pct = int(current / total * 100) if total else 0
    bar = ("█" * (pct // 5)).ljust(20)
    print(f"\r  [{bar}] {pct:3d}%  {current}/{total}  {label}      ", end="", flush=True)
    if current == total:
        print()


# ---------------------------------------------------------------------------
# API Client with rate-limit tracking
# ---------------------------------------------------------------------------
class MistAPI:
    def __init__(self, base_url: str, token: str):
        self.base   = base_url.rstrip("/")
        self.token  = token
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Token {token}",
            "Content-Type":  "application/json",
        })
        self._call_count   = 0
        self._window_start = time.time()

    # ---- rate-limit guard --------------------------------------------------
    def _check_rate_limit(self):
        now = time.time()
        elapsed = now - self._window_start
        if elapsed >= API_CALL_WINDOW:
            self._call_count   = 0
            self._window_start = now
        if self._call_count >= (API_RATE_LIMIT - RATE_LIMIT_BUFFER):
            wait = API_CALL_WINDOW - elapsed + 5
            log(f"Rate-limit guard: {self._call_count} calls made. Sleeping {wait:.0f}s …", "WARN")
            time.sleep(wait)
            self._call_count   = 0
            self._window_start = time.time()

    # ---- low-level request -------------------------------------------------
    def _request(self, method, path, payload=None, params=None):
        self._check_rate_limit()
        url = f"{self.base}/api/v1{path}"
        debug(f"{method} {url}  params={params}")
        try:
            r = self.session.request(method, url, json=payload, params=params, timeout=30)
            self._call_count += 1
            debug(f"  → {r.status_code}  calls_this_window={self._call_count}")
            if r.status_code == 429:
                retry_after = int(r.headers.get("Retry-After", 60))
                log(f"429 Too Many Requests – waiting {retry_after}s", "WARN")
                time.sleep(retry_after)
                return self._request(method, path, payload, params)
            r.raise_for_status()
            return r.json() if r.text else {}
        except requests.exceptions.RequestException as e:
            log(f"API ERROR: {method} {path} → {e}", "ERROR")
            debug(f"  Exception detail: {e}")
            return None

    def get(self, path, params=None):
        return self._request("GET", path, params=params)

    def put(self, path, payload):
        return self._request("PUT", path, payload=payload)

    def delete(self, path):
        return self._request("DELETE", path)

    @property
    def call_count(self):
        return self._call_count

    # ---- paginated GET helper ----------------------------------------------
    def get_all(self, path, params=None, key=None):
        """Fetch all pages of a paginated endpoint."""
        results  = []
        page     = 1
        per_page = 100
        p        = dict(params or {})
        p.update({"page": page, "limit": per_page})
        while True:
            p["page"] = page
            data = self.get(path, params=p)
            if data is None:
                break
            chunk = data if isinstance(data, list) else data.get(key or "results", data)
            if not isinstance(chunk, list):
                # some endpoints return list directly
                chunk = data if isinstance(data, list) else []
            results.extend(chunk)
            if len(chunk) < per_page:
                break
            page += 1
        return results


# ---------------------------------------------------------------------------
# Cloud & Auth prompts
# ---------------------------------------------------------------------------
def prompt_cloud() -> str:
    section("Mist Cloud Selection")
    for k, v in MIST_CLOUDS.items():
        print(f"  {k}) {v['name']}  ({v['base']})")
    while True:
        choice = input("\n  Enter cloud number [1-6]: ").strip()
        if choice in MIST_CLOUDS:
            base = MIST_CLOUDS[choice]["base"]
            log(f"Selected cloud: {MIST_CLOUDS[choice]['name']}  →  {base}")
            return base
        print("  Invalid choice – try again.")

def prompt_org_id() -> str:
    section("Org ID")
    while True:
        org = input("  Paste your Mist Org ID (UUID): ").strip()
        if re.match(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", org, re.I):
            log(f"Org ID: {org}")
            return org
        print("  Does not look like a valid UUID – try again.")

def prompt_token(label="Read-Only API Token") -> str:
    print(f"\n  {label} (input hidden):")
    token = getpass.getpass("  Token: ").strip()
    if not token:
        sys.exit("ERROR: Token cannot be empty.")
    return token

def verify_auth(api: MistAPI) -> dict:
    """Call /self to confirm token is valid and return user info."""
    data = api.get("/self")
    if not data:
        sys.exit("ERROR: Authentication failed – check token and cloud selection.")
    log(f"Authenticated as: {data.get('email', 'unknown')}  (privileges: {len(data.get('privileges', []))})", "OK")
    return data


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------
def get_sites(api: MistAPI, org_id: str) -> list:
    log("Fetching sites …")
    sites = api.get(f"/orgs/{org_id}/sites") or []
    log(f"  Found {len(sites)} site(s).", "OK")
    return sites

def get_wlan_templates(api: MistAPI, org_id: str) -> list:
    log("Fetching WLAN templates …")
    templates = api.get(f"/orgs/{org_id}/wlantemplates") or []
    log(f"  Found {len(templates)} WLAN template(s).", "OK")
    return templates

def get_template_detail(api: MistAPI, org_id: str, tmpl_id: str) -> dict:
    return api.get(f"/orgs/{org_id}/wlantemplates/{tmpl_id}") or {}

def get_site_wlans(api: MistAPI, site_id: str) -> list:
    return api.get(f"/sites/{site_id}/wlans") or []

def get_org_client_count(api: MistAPI, org_id: str) -> int:
    """Return total connected wireless clients across the org."""
    data = api.get(f"/orgs/{org_id}/stats/clients", params={"limit": 1})
    if isinstance(data, list):
        # some versions return list directly; full count via header not available
        # use a reasonable page approach
        all_clients = api.get_all(f"/orgs/{org_id}/stats/clients")
        return len(all_clients)
    if isinstance(data, dict):
        return data.get("total", 0)
    return 0

def get_site_client_count(api: MistAPI, site_id: str) -> int:
    clients = api.get(f"/sites/{site_id}/stats/clients") or []
    return len(clients) if isinstance(clients, list) else 0

def get_sle_successful_connect(api: MistAPI, site_id: str, hours=24) -> dict:
    """
    Fetch Successful Connect SLE summary for a site.
    Returns dict with 'value' (percentage 0-100) or empty dict on failure.
    """
    end_ts   = int(time.time())
    start_ts = end_ts - (hours * 3600)
    path     = f"/sites/{site_id}/sle/wireless/metric/successful-connect/summary"
    params   = {"start": start_ts, "end": end_ts}
    data = api.get(path, params=params)
    if data is None:
        return {}
    return data


# ---------------------------------------------------------------------------
# Best Practices definitions
# ---------------------------------------------------------------------------
BEST_PRACTICES = [
    {
        "id":       "arp_filter",
        "name":     "ARP Filtering",
        "field":    "arp_filter",
        "desired":  True,
        "impact":   (
            "ARP broadcast storms can degrade Wi-Fi performance in dense environments. "
            "Enabling ARP filtering causes the AP to respond to ARP requests on behalf of "
            "known clients (proxy-ARP), dramatically reducing broadcast traffic over the air. "
            "This is especially impactful in large-scale deployments and IoT-heavy networks."
        ),
        "remediation_key": "arp_filter",
    },
    {
        "id":       "limit_bcast",
        "name":     "Multicast/Broadcast Filtering",
        "field":    "limit_bcast",
        "desired":  True,
        "impact":   (
            "Broadcast and multicast frames are sent at the lowest basic data rate and must be "
            "received by every client, consuming valuable airtime. Enabling broadcast/multicast "
            "filtering suppresses unnecessary layer-2 floods, reducing channel utilisation and "
            "improving overall throughput and latency for all clients on that SSID."
        ),
        "remediation_key": "limit_bcast",
    },
    {
        "id":       "allow_ipv6_ndp",
        "name":     "Allow IPv6 Clients",
        "field":    "allow_ipv6_ndp",
        "desired":  True,
        "impact":   (
            "IPv6 Neighbor Discovery Protocol (NDP) replaces ARP in IPv6 networks. If this is "
            "disabled, dual-stack and IPv6-only clients cannot perform address resolution and "
            "will fail to communicate. Enabling NDP ensures forward compatibility with modern "
            "operating systems that prefer IPv6, and avoids silent connectivity failures."
        ),
        "remediation_key": "allow_ipv6_ndp",
    },
    {
        "id":       "data_rates",
        "name":     "Data Rates (no 802.11b legacy)",
        "field":    None,       # derived check – see check_data_rates()
        "desired":  True,
        "impact":   (
            "Legacy 802.11b rates (1, 2, 5.5, 11 Mbps) were designed for early Wi-Fi and are "
            "rarely needed today. When enabled they force the AP to transmit beacons and "
            "management frames at the lowest supported rate, wasting airtime and reducing "
            "capacity for all modern clients. Disabling 802.11b rates (setting minimum to "
            "6 Mbps or higher) significantly improves network efficiency."
        ),
        "remediation_key": None,  # Managed at RF-template level in Mist
    },
    {
        "id":       "dot11r",
        "name":     "802.11r Fast Transition (Enterprise WPA2/WPA3)",
        "field":    None,       # derived check – see check_dot11r()
        "desired":  True,
        "impact":   (
            "802.11r Fast BSS Transition (FT) enables clients to pre-authenticate with "
            "neighboring APs before roaming, reducing roam latency from >100 ms to <50 ms. "
            "This is critical for voice, video, and real-time applications. Mist supports "
            "Hybrid/Mixed 802.11r so legacy clients that do not support FT are unaffected. "
            "Only applicable to WPA2/WPA3 Enterprise SSIDs."
        ),
        "remediation_key": None,  # nested in auth object
    },
    {
        "id":       "duplicate_ssid",
        "name":     "No Duplicate WLAN Names per Site",
        "field":    None,       # derived – see check_duplicate_ssids()
        "desired":  True,
        "impact":   (
            "Duplicate SSID names on the same site can cause unexpected client behaviour, "
            "authentication loops, and makes troubleshooting significantly harder. Each SSID "
            "should have a unique name within a site unless there is an intentional reason "
            "(e.g. 2.4/5 GHz split – which Mist handles automatically). Duplicates are "
            "usually the result of a misconfiguration."
        ),
        "remediation_key": None,  # Must be manually resolved
    },
]


# ---------------------------------------------------------------------------
# Best practice checkers
# ---------------------------------------------------------------------------
def check_data_rates(wlan: dict) -> bool:
    """
    Return True (compliant) if the WLAN has no 802.11b legacy rates.
    Mist WLAN data_rates field – 'legacy' array lists permitted legacy rates.
    Compliant = field absent OR legacy list does not contain 11b rates.
    Note: Primary data-rate control is in the RF Template in Mist.
    """
    dr = wlan.get("data_rates")
    if dr is None:
        # Not configured at WLAN level; check is at RF template – mark as INFO
        return None   # None = not applicable / check at RF template
    legacy = dr.get("legacy", [])
    eleven_b = {"1", "2", "5.5", "11", "11b", "1m", "2m", "5.5m", "11m"}
    return not bool(set(str(r) for r in legacy) & eleven_b)

def check_dot11r(wlan: dict) -> bool:
    """Return True if 802.11r is enabled for enterprise auth WLANs."""
    auth = wlan.get("auth", {})
    auth_type = auth.get("type", "open").lower()
    if auth_type not in ENTERPRISE_AUTH_TYPES:
        return None  # Not applicable (not enterprise)
    disable_ft = auth.get("disable_ft", True)
    return not disable_ft   # disable_ft=False means FT is ENABLED → compliant

def check_duplicate_ssids(wlans: list) -> list:
    """Return list of SSIDs that appear more than once in the wlan list."""
    seen   = defaultdict(int)
    for w in wlans:
        seen[w.get("ssid", "")] += 1
    return [ssid for ssid, cnt in seen.items() if cnt > 1]

def evaluate_wlan(wlan: dict) -> dict:
    """
    Evaluate a single WLAN against all applicable best practices.
    Returns dict keyed by BP id → {'compliant': bool|None, 'current': value}
    """
    results = {}
    for bp in BEST_PRACTICES:
        bid = bp["id"]
        if bid == "data_rates":
            c   = check_data_rates(wlan)
            cur = wlan.get("data_rates", "not set at WLAN level")
        elif bid == "dot11r":
            c   = check_dot11r(wlan)
            cur = wlan.get("auth", {}).get("disable_ft", "N/A")
        elif bid == "duplicate_ssid":
            continue  # checked at site level
        else:
            field = bp["field"]
            cur   = wlan.get(field)
            c     = (cur == bp["desired"]) if cur is not None else False
        results[bid] = {"compliant": c, "current": cur}
    return results


# ---------------------------------------------------------------------------
# Display helpers
# ---------------------------------------------------------------------------
PASS  = "\033[92m✔ PASS\033[0m"
FAIL  = "\033[91m✘ FAIL\033[0m"
NA    = "\033[93m– N/A \033[0m"
INFO_ = "\033[94mℹ INFO\033[0m"

def status_str(val):
    if val is True:   return PASS
    if val is False:  return FAIL
    return NA

def print_bp_table(tmpl_name: str, wlan_name: str, bp_results: dict):
    print(f"\n  WLAN: \033[1m{wlan_name}\033[0m  (template: {tmpl_name})")
    print(f"  {'Best Practice':<45} {'Status':<12} {'Current Value'}")
    print(f"  {'-'*45} {'-'*12} {'-'*30}")
    for bp in BEST_PRACTICES:
        bid  = bp["id"]
        if bid == "duplicate_ssid":
            continue
        r    = bp_results.get(bid, {})
        comp = r.get("compliant")
        cur  = str(r.get("current", ""))[:40]
        print(f"  {bp['name']:<45} {status_str(comp):<20} {cur}")

def print_best_practices_guide():
    section("WLAN Best Practices Reference Guide")
    for i, bp in enumerate(BEST_PRACTICES, 1):
        print(f"\n  {i}. \033[1m{bp['name']}\033[0m")
        # wrap at ~72 chars
        words  = bp["impact"].split()
        line   = "     "
        for w in words:
            if len(line) + len(w) + 1 > 72:
                print(line)
                line = "     " + w + " "
            else:
                line += w + " "
        if line.strip():
            print(line)


# ---------------------------------------------------------------------------
# Collect all data
# ---------------------------------------------------------------------------
def collect_all(api: MistAPI, org_id: str):
    """
    Returns a rich dict of all collected data.
    """
    data = {
        "org_id":       org_id,
        "collected_at": _ts(),
        "sites":        [],
        "templates":    [],
        "site_wlans":   {},     # site_id → list of wlan dicts
        "template_wlans": {},   # template_id → list of wlan dicts
        "clients_before": {},   # site_id → count
        "clients_total_before": 0,
        "sle_before":   {},     # site_id → sle dict
    }

    sites     = get_sites(api, org_id)
    templates = get_wlan_templates(api, org_id)
    data["sites"]     = sites
    data["templates"] = templates

    # Detailed template WLANs
    section("Loading WLAN Template Details")
    for i, tmpl in enumerate(templates):
        tid  = tmpl["id"]
        progress(i + 1, len(templates), tmpl.get("name", tid))
        detail = get_template_detail(api, org_id, tid)
        data["template_wlans"][tid] = detail.get("wlans", [])

    # Site WLANs + clients + SLE
    section("Loading Site WLANs, Client Counts, and SLE")
    for i, site in enumerate(sites):
        sid  = site["id"]
        sname = site.get("name", sid)
        progress(i + 1, len(sites), sname)
        data["site_wlans"][sid]   = get_site_wlans(api, sid)
        data["clients_before"][sid] = get_site_client_count(api, sid)
        sle = get_sle_successful_connect(api, sid, hours=24)
        data["sle_before"][sid] = sle

    # Total client count
    total = sum(data["clients_before"].values())
    data["clients_total_before"] = total

    log(f"  Total wireless clients (org): {total}", "OK")
    return data


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------
def report_client_summary(data: dict):
    section("WLAN Client Summary")
    total = data["clients_total_before"]
    log(f"  Total wireless clients across org: {total}", "OK")
    print(f"\n  {'Site':<40} {'Clients':>8}")
    print(f"  {'-'*40} {'-'*8}")
    for site in data["sites"]:
        sid   = site["id"]
        cnt   = data["clients_before"].get(sid, 0)
        print(f"  {site.get('name', sid):<40} {cnt:>8}")

def report_sle_summary(data: dict):
    section("Successful Connect SLE – Last 24 Hours")
    print(f"\n  {'Site':<40} {'SLE Value':>12}")
    print(f"  {'-'*40} {'-'*12}")
    for site in data["sites"]:
        sid  = site["id"]
        sle  = data["sle_before"].get(sid, {})
        val  = sle.get("value") or sle.get("sle_value") or sle.get("avg") or "N/A"
        if isinstance(val, (int, float)):
            val = f"{val:.1f}%"
        print(f"  {site.get('name', sid):<40} {str(val):>12}")

def report_bp_status(data: dict):
    section("WLAN Best Practices Status by Template")
    all_results = []  # for Excel export
    for tmpl in data["templates"]:
        tid   = tmpl["id"]
        tname = tmpl.get("name", tid)
        wlans = data["template_wlans"].get(tid, [])
        print(f"\n  \033[1mTemplate: {tname}\033[0m")
        if not wlans:
            print("    (no WLANs in this template)")
            continue
        for wlan in wlans:
            bp_res = evaluate_wlan(wlan)
            print_bp_table(tname, wlan.get("ssid", "?"), bp_res)
            all_results.append({
                "template_id":   tid,
                "template_name": tname,
                "wlan_id":       wlan.get("id", ""),
                "ssid":          wlan.get("ssid", ""),
                "bp_results":    bp_res,
                "wlan_obj":      wlan,
            })
    return all_results

def report_site_wlans(data: dict):
    section("Site WLAN Listing")
    for site in data["sites"]:
        sid   = site["id"]
        sname = site.get("name", sid)
        wlans = data["site_wlans"].get(sid, [])
        print(f"\n  Site: \033[1m{sname}\033[0m  ({len(wlans)} WLAN{'s' if len(wlans)!=1 else ''})")
        for w in wlans:
            print(f"    - {w.get('ssid','?')}  (id: {w.get('id','')})")

def report_duplicate_ssids(data: dict):
    section("Duplicate WLAN Names per Site")
    found_any = False
    for site in data["sites"]:
        sid   = site["id"]
        wlans = data["site_wlans"].get(sid, [])
        dupes = check_duplicate_ssids(wlans)
        if dupes:
            found_any = True
            log(f"  Site \"{site.get('name',sid)}\" has duplicate SSIDs: {dupes}", "WARN")
    if not found_any:
        log("  No duplicate SSID names found across any site.", "OK")


# ---------------------------------------------------------------------------
# Interactive flow
# ---------------------------------------------------------------------------
def ask_yn(prompt_text: str) -> bool:
    while True:
        r = input(f"\n  {prompt_text} [y/N]: ").strip().lower()
        if r in ("y", "yes"):
            return True
        if r in ("n", "no", ""):
            return False
        print("  Please enter y or n.")

def interactive_site_wlan_menu(api: MistAPI, data: dict):
    if not ask_yn("Do you want to list each Site and the WLANs present?"):
        return
    report_site_wlans(data)

    if ask_yn("Do you want to delete any WLAN by site?"):
        rw_token  = prompt_token("Read/Write API Token")
        rw_api    = MistAPI(api.base, rw_token)
        verify_auth(rw_api)

        print("\n  WARNING: \033[91mDeleting a WLAN from a site may bounce radios and cause a brief wireless outage!\033[0m")
        if not ask_yn("Are you SURE you want to proceed with WLAN deletion?"):
            log("WLAN deletion cancelled by user.", "WARN")
            return

        site_name = input("  Enter the exact Site Name to delete from: ").strip()
        site_obj  = next((s for s in data["sites"] if s.get("name") == site_name), None)
        if not site_obj:
            log(f"  Site '{site_name}' not found.", "ERROR")
            return

        ssid_name = input("  Enter the exact SSID name to delete: ").strip()
        site_wlans = data["site_wlans"].get(site_obj["id"], [])
        wlan_obj  = next((w for w in site_wlans if w.get("ssid") == ssid_name), None)
        if not wlan_obj:
            log(f"  SSID '{ssid_name}' not found on site '{site_name}'.", "ERROR")
            return

        wlan_id = wlan_obj["id"]
        log(f"  Deleting WLAN '{ssid_name}' (id: {wlan_id}) from site '{site_name}' …", "WARN")
        result = rw_api.delete(f"/sites/{site_obj['id']}/wlans/{wlan_id}")
        if result is not None:
            log(f"  WLAN '{ssid_name}' deleted successfully.", "OK")
        else:
            log(f"  Deletion failed – check debug log.", "ERROR")

    if ask_yn("Do you want to show sites with duplicate WLAN names?"):
        report_duplicate_ssids(data)


def apply_best_practices(api_rw: MistAPI, org_id: str, data: dict, bp_results: list):
    """
    Prompt user to enable best practices per WLAN template.
    Returns list of change records.
    """
    section("Apply WLAN Best Practices")
    changes = []

    for tmpl in data["templates"]:
        tid   = tmpl["id"]
        tname = tmpl.get("name", tid)
        wlans = data["template_wlans"].get(tid, [])
        if not wlans:
            continue

        # Find non-compliant items
        non_compliant = []
        for wlan in wlans:
            bp_r = evaluate_wlan(wlan)
            for bid, res in bp_r.items():
                if res.get("compliant") is False:
                    non_compliant.append((wlan, bid))

        if not non_compliant:
            log(f"  Template '{tname}': all checked items compliant – skipping.", "OK")
            continue

        print(f"\n  Template: \033[1m{tname}\033[0m – {len(non_compliant)} non-compliant item(s) found.")

        for wlan, bid in non_compliant:
            ssid = wlan.get("ssid", "?")
            bp   = next(b for b in BEST_PRACTICES if b["id"] == bid)
            print(f"\n    WLAN: {ssid}  |  BP: {bp['name']}")
            if not ask_yn(f"    Enable '{bp['name']}' on WLAN '{ssid}'?"):
                log(f"  Skipped '{bp['name']}' on {ssid}", "WARN")
                continue

            # Build patch payload
            payload = dict(wlan)   # start with existing config
            rem_key = bp.get("remediation_key")
            if rem_key:
                payload[rem_key] = True
            elif bid == "dot11r":
                auth = dict(payload.get("auth", {}))
                auth["disable_ft"] = False
                payload["auth"] = auth
            elif bid == "data_rates":
                log("  Data rates are managed at the RF Template level in Mist – cannot patch via WLAN API.", "WARN")
                log("  Please update the RF Template minimum data rate to 6 Mbps or higher in the Mist UI.", "WARN")
                continue

            wlan_id = wlan.get("id")
            log(f"  Patching WLAN '{ssid}' (id: {wlan_id}) in template '{tname}' …")
            result = api_rw.put(f"/orgs/{org_id}/wlantemplates/{tid}/wlans/{wlan_id}", payload)
            if result:
                log(f"  ✔ '{bp['name']}' enabled on WLAN '{ssid}'", "OK")
                changes.append({"template": tname, "ssid": ssid, "bp": bp["name"],
                                 "action": "enabled", "status": "success"})
            else:
                log(f"  ✘ Failed to patch WLAN '{ssid}' – see debug log", "ERROR")
                changes.append({"template": tname, "ssid": ssid, "bp": bp["name"],
                                 "action": "enabled", "status": "failed"})
    return changes


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------
def export_excel(data: dict, bp_results: list, output_dir: str):
    if not XLSX_AVAILABLE:
        log("Excel export skipped – openpyxl not installed.", "WARN")
        return None

    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"mist_bp_report_{ts}.xlsx")
    wb   = openpyxl.Workbook()

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    na_fill   = PatternFill("solid", fgColor="FFEB9C")
    thin      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    def style_header(cell):
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    def style_data(cell, fill=None):
        if fill:
            cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True)

    bp_ids   = [b["id"] for b in BEST_PRACTICES]
    bp_names = [b["name"] for b in BEST_PRACTICES]

    # ---- Tab 1: By WLAN Template ----------------------------------------
    ws1 = wb.active
    ws1.title = "By WLAN Template"
    headers = ["Template Name", "SSID"] + bp_names + ["Auth Type"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        style_header(c)

    row = 2
    for r in bp_results:
        ws1.cell(row=row, column=1, value=r["template_name"])
        ws1.cell(row=row, column=2, value=r["ssid"])
        for ci, bid in enumerate(bp_ids, 3):
            comp = r["bp_results"].get(bid, {}).get("compliant")
            val  = "PASS" if comp is True else ("FAIL" if comp is False else "N/A")
            fill = pass_fill if comp is True else (fail_fill if comp is False else na_fill)
            c = ws1.cell(row=row, column=ci, value=val)
            style_data(c, fill)
        # auth type
        auth_type = r["wlan_obj"].get("auth", {}).get("type", "open")
        ws1.cell(row=row, column=len(headers), value=auth_type)
        row += 1

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 22
    ws1.freeze_panes = "A2"

    # ---- Tab 2: By Site --------------------------------------------------
    ws2 = wb.create_sheet("By Site")
    s_headers = ["Site Name", "SSID", "Inherited Template"] + bp_names
    for col, h in enumerate(s_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        style_header(c)

    row2 = 2
    # Build site→template map
    site_tmpl = {}
    for site in data["sites"]:
        site_tmpl[site["id"]] = site.get("wlantemplate_id", "")

    for site in data["sites"]:
        sid   = site["id"]
        sname = site.get("name", sid)
        wlans = data["site_wlans"].get(sid, [])
        for wlan in wlans:
            bp_r = evaluate_wlan(wlan)
            ws2.cell(row=row2, column=1, value=sname)
            ws2.cell(row=row2, column=2, value=wlan.get("ssid", ""))
            ws2.cell(row=row2, column=3, value=site_tmpl.get(sid, ""))
            for ci, bid in enumerate(bp_ids, 4):
                comp = bp_r.get(bid, {}).get("compliant")
                val  = "PASS" if comp is True else ("FAIL" if comp is False else "N/A")
                fill = pass_fill if comp is True else (fail_fill if comp is False else na_fill)
                c = ws2.cell(row=row2, column=ci, value=val)
                style_data(c, fill)
            row2 += 1

    for col in range(1, len(s_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    ws2.freeze_panes = "A2"

    # ---- Tab 3: By AP (site-level client proxy) --------------------------
    ws3 = wb.create_sheet("By AP")
    ap_headers = ["Site Name", "Total Clients (site)", "SLE Successful Connect (24h)"]
    for col, h in enumerate(ap_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        style_header(c)

    row3 = 2
    for site in data["sites"]:
        sid  = site["id"]
        sle  = data["sle_before"].get(sid, {})
        val  = sle.get("value") or sle.get("sle_value") or sle.get("avg") or "N/A"
        if isinstance(val, (int, float)):
            val = f"{val:.1f}%"
        ws3.cell(row=row3, column=1, value=site.get("name", sid))
        ws3.cell(row=row3, column=2, value=data["clients_before"].get(sid, 0))
        ws3.cell(row=row3, column=3, value=str(val))
        row3 += 1

    for col in range(1, len(ap_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 30
    ws3.freeze_panes = "A2"

    wb.save(path)
    log(f"  Excel report saved → {path}", "OK")
    return path


# ---------------------------------------------------------------------------
# Post-change verification
# ---------------------------------------------------------------------------
def post_change_verification(api: MistAPI, org_id: str, data: dict):
    section("Post-Change Verification")
    log("  Waiting 30 seconds for changes to propagate …")
    time.sleep(30)

    log("  Checking client counts after changes …")
    total_after = 0
    for site in data["sites"]:
        sid  = site["id"]
        cnt  = get_site_client_count(api, sid)
        before = data["clients_before"].get(sid, 0)
        total_after += cnt
        delta = cnt - before
        sym   = "▲" if delta > 0 else ("▼" if delta < 0 else "=")
        log(f"  {site.get('name', sid):<40} before={before}  after={cnt}  {sym}{abs(delta)}")

    log(f"  Org-wide clients: before={data['clients_total_before']}  after={total_after}")

    log("  Checking SLE Successful Connect for past 60 minutes …")
    for site in data["sites"]:
        sid  = site["id"]
        sle_after = get_sle_successful_connect(api, sid, hours=1)
        val_after = sle_after.get("value") or sle_after.get("sle_value") or "N/A"
        val_before_raw = data["sle_before"].get(sid, {})
        val_before = val_before_raw.get("value") or val_before_raw.get("sle_value") or "N/A"
        if isinstance(val_after, (int, float)):
            val_after = f"{val_after:.1f}%"
        if isinstance(val_before, (int, float)):
            val_before = f"{val_before:.1f}%"
        log(f"  {site.get('name', sid):<40} SLE before (24h)={val_before}  SLE after (1h)={val_after}")


# ---------------------------------------------------------------------------
# Scheduled / midnight automation hook
# ---------------------------------------------------------------------------
def schedule_midnight_run():
    """
    Inform user how to set up a cron job for midnight automation.
    The actual scheduling uses system cron or Windows Task Scheduler.
    """
    section("Midnight Automation Setup")
    script_path = os.path.abspath(__file__)
    log("  To run this script automatically every midnight, add the following cron entry:")
    log(f"  0 0 * * *  /usr/bin/python3 {script_path} --auto")
    log("  On Windows, use Task Scheduler to run:")
    log(f"  python \"{script_path}\" --auto")
    log("  '--auto' flag skips interactive prompts and uses environment variables:")
    log("    MIST_CLOUD=1  MIST_ORG_ID=<uuid>  MIST_TOKEN=<token>")
    log("  Save token to environment: export MIST_TOKEN=your_token_here")


# ---------------------------------------------------------------------------
# Auto mode (non-interactive)
# ---------------------------------------------------------------------------
def run_auto_mode(base_url: str, org_id: str, token: str):
    """
    Non-interactive mode used for scheduled midnight runs.
    Applies all best practices automatically and logs results.
    """
    api = MistAPI(base_url, token)
    verify_auth(api)

    data      = collect_all(api, org_id)
    bp_results = report_bp_status(data)

    # Apply with a rw token (same token assumed to have rw access in auto mode)
    changes   = apply_best_practices(api, org_id, data, bp_results)
    log(f"  Auto-mode: {len(changes)} change(s) applied.", "OK")

    post_change_verification(api, org_id, data)

    output_dir = os.path.dirname(os.path.abspath(__file__))
    export_excel(data, bp_results, output_dir)
    log(f"  API calls this session: {api.call_count}")
    log(f"  Total elapsed time: {_elapsed()}")
    log(f"  Log file: {LOG_FILE}")
    log(f"  Debug log: {DEBUG_LOG}")


# ---------------------------------------------------------------------------
# Main interactive flow
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Mist WLAN Best Practices Automation")
    parser.add_argument("--auto", action="store_true",
                        help="Non-interactive mode (reads MIST_CLOUD, MIST_ORG_ID, MIST_TOKEN env vars)")
    args = parser.parse_args()

    section(f"Juniper Mist WLAN Best Practices Automation  v{SCRIPT_VERSION}")
    log(f"  Start time : {_ts()}")
    log(f"  Log file   : {LOG_FILE}")
    log(f"  Debug log  : {DEBUG_LOG}")

    # --- Auto mode -----------------------------------------------------------
    if args.auto:
        cloud_key = os.environ.get("MIST_CLOUD", "1")
        cloud_obj = MIST_CLOUDS.get(cloud_key, MIST_CLOUDS["1"])
        org_id    = os.environ.get("MIST_ORG_ID", "")
        token     = os.environ.get("MIST_TOKEN", "")
        if not org_id or not token:
            sys.exit("ERROR: MIST_ORG_ID and MIST_TOKEN environment variables required in auto mode.")
        run_auto_mode(cloud_obj["base"], org_id, token)
        return

    # --- Interactive mode ----------------------------------------------------
    base_url = prompt_cloud()
    org_id   = prompt_org_id()
    ro_token = prompt_token("Read-Only API Token")

    api = MistAPI(base_url, ro_token)
    verify_auth(api)

    # Collect all data
    data = collect_all(api, org_id)

    # Summaries
    report_client_summary(data)
    report_sle_summary(data)

    # Best practices guide
    if ask_yn("Do you want to view the WLAN Best Practices reference guide?"):
        print_best_practices_guide()

    # BP status
    section("Checking WLAN Best Practices Status")
    bp_results = report_bp_status(data)

    # Interactive site/WLAN menu
    interactive_site_wlan_menu(api, data)

    # Duplicate SSIDs
    report_duplicate_ssids(data)

    # Offer to apply best practices
    if ask_yn("Do you want to enable the applicable WLAN Best Practices now?"):
        rw_token  = prompt_token("SuperUser / Read-Write API Token")
        rw_api    = MistAPI(base_url, rw_token)
        verify_auth(rw_api)
        changes = apply_best_practices(rw_api, org_id, data, bp_results)
        log(f"  {len(changes)} change(s) applied.", "OK")
        post_change_verification(api, org_id, data)
    else:
        changes = []

    # Midnight automation
    if ask_yn("Do you want to automatically check and apply WLAN Best Practices at midnight each night?"):
        schedule_midnight_run()

    # Excel export
    output_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = export_excel(data, bp_results, output_dir)

    # Final summary
    section("Run Summary")
    log(f"  Sites checked         : {len(data['sites'])}")
    log(f"  Templates checked     : {len(data['templates'])}")
    log(f"  Total wireless clients: {data['clients_total_before']}")
    log(f"  Changes applied       : {len(changes)}")
    log(f"  API calls this session: {api.call_count}")
    log(f"  Total elapsed time    : {_elapsed()}")
    log(f"  Main log file         : {LOG_FILE}")
    log(f"  Debug log file        : {DEBUG_LOG}")
    if excel_path:
        log(f"  Excel report          : {excel_path}", "OK")


if __name__ == "__main__":
    main()
